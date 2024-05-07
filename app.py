# pip install openpyxl

import priceCrawling
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from datetime import datetime

print("JSON PARSING start")

#닷컴 크롤링 값 - 컨피규 json값 비교
carList = ['Sportage', 'K8', 'Sorento', 'EV6', 'EV9', 'Carnival']
fileName=['스포티지HEV', 'K8', '쏘렌토', 'EV6', 'EV9', '카니발']
taxList = ['일반인', '다자녀', '장애 4~6급', '국가유공자', '광주민주화']
fpath = os.getcwd()
taxN = 7 # 가격 비교 종류 수


def read_Json(fpath, car, tax, prefix):
    fName=prefix+'_'+tax
    fpath += '/JsonData/'+car+'/'+fName+'.json'
    #없는 파일에 대해 에러처리
    #json데이터
    res = []
    with open(fpath) as json_file:
        json_data = json.load(json_file)
        tot = json_data['quoteTotalAmount'] #결제 금액 D
        aqT = json_data['acquisitionTax'] # 취득세 F
        bond = json_data['bond'] # 공채 G
        licenseP = json_data['carLicensePlate'] #차량 번호판 H
        reg = json_data['registrationProxyCarge'] # 등록 대행 I
        iT = json_data['individualTax'] # 세금 J
        totT = str(int(licenseP)+int(reg)+int(aqT)+int(bond))
        res = [tot, totT, aqT, bond, licenseP, reg, iT]

    return res


#엑셀 파일 읽기
currentTime=str(datetime.today().strftime("%Y-%m-%d")) 
p = os.getcwd()+"/result/KiaPrice_"+currentTime+".xlsx"
wb=load_workbook(p, data_only=True)
ws = wb.active

confList = []

#차량기준 비교
idx = 2 #처음 2행부터 시작
for car in carList:
    for tax in taxList:
        kia = [] #크롤링한 가격 정보
        if ws[idx][0].value != car:
            print(ws[idx][0].value, car, "엑셀 차량 순서 확인")
            exit()

        for col in ws[idx][3:3+taxN]:
            kia.append(col.value.replace(',', ''))
        # print("kia : " ,kia)
        
        #차별 Json 읽기
        conf = read_Json(fpath, car, tax, fileName) # json에서 읽은 가격정보
        # print("conf : ", conf)

        #정합성 체크, 엑셀쓰기
        if len(kia) != len(conf):
            print("엑셀-json 가격 정보 부족")
            exit()
        
        y = 11 #K열부터 작성
        for i in range(taxN):
            # print(conf[i])
            ws.cell(row=idx, column=y+i).value = conf[i]
            
            if kia[i] != conf[i]:
                ws.cell(row=idx, column=y+i).fill = PatternFill(fill_type='solid', fgColor=Color('789ABC'))

        idx +=1
    
print("print result============")
wb.save("result/result_" + currentTime+ ".xlsx") #경로 


            